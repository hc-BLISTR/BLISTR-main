# -*- coding: utf-8 -*-
"""
Created on Tue Jul 26 15:40:24 2016

@author: Owen Neuber

Program to produce an excel table which will be converted into an html table.
It will be based off of 'metadata_table.xlsx's first 3 rows.
This program is UNUSED in BLISTR
"""


import xlsxwriter
from openpyxl import load_workbook


workbook_path = 'PATH/TO/ROOT/Database_Inputs/'
wb = load_workbook(filename=workbook_path+'metadata_table.xlsx', read_only=True)
ws = wb['Sheet1'] # ws is now an IterableWorksheet

#Output workbook
#override the old metadata_table.xlsx file with the new one!    
workbook = xlsxwriter.Workbook(workbook_path+'toxin_table.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': 1}) #allowing bold wording because why not

temp_g_name = "" #initializing a variable to determine if a new line is needed or is the last line should be replaced

headers = [" ", "Genome Name", "Collection date"] #the blank represents check boxes
#Copy the old workbook out exactly as it was...
counter = 0
row_num = 1
letter = 'a'
header_3 = 0
while header_3 < 4:
    for header in headers: #we want this to be run at the top 4 times            
        worksheet.write(letter.upper()+'1', str(header), bold) #sets the cell postion (i.e. A1, B1, C1 etc.) and makes the header bold
        letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter]) #wizard one linner that will increase the letter value of letter (i.e 'a' -> 'b')
    header_3 += 1
for row in ws.rows:
    if letter == 'm':
        letter = 'a' #if we hit the third run, restart at the right side of the table
        row_num += 1 #only increase row number when we have finished all our inputs for the row
    counter += 1
    column_num = 0 #if this is 1, there is significance to temp_g_name
    for cell in row:
        if column_num > 2:
            break #I only want the first 3 lines
        if counter == 1: #for the header row, we want to copy it as is
            continue
        else:
            worksheet.write(letter.upper()+str(row_num), str(cell.value))
            letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter])
            if column_num == 1:
                temp_g_name = str(cell.value) #keep track of the most recent genome name
        column_num += 1



workbook.close() #close whatever workbook we may have been working on                
