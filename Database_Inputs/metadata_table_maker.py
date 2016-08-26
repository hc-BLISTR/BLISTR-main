# -*- coding: utf-8 -*-
"""
Created on Fri Jul  8 14:04:38 2016

@author: Owen Neuber

This program will take all the information in the database and print the metadata 
to a .xlsx file in the directory: PATH/TO/ROOT/Database_Inputs/
That file will be used to display information on the Genomes page. Additionally,
if there is to be additional genomes added to the database, this program will
update this metadata file to hold this new information (rather than create the
whole table from scrach again)

The outputted file will be nammed: metadata_table.xlsx

This program will be run every time a genome is uploaded to the database (via processor.py),
when PROKKA finishes running (called in PROKKA.py) and also everytime db_loaders.py (in 
Database_Crseation) run. When db_loaders is run, the -n flag will be given to create the 
table from scratch. When processor.py runs it, no flag will be given and a single row will 
be added. Should PROKKA.py run, no flags will be given and one row will be modified
"""

import xlsxwriter
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine
from models import Genome, GeographicLocation, Toxin, Prokka#, FlaA, FlaB
from openpyxl import load_workbook
from tqdm import tqdm
import argparse
parser = argparse.ArgumentParser()
parser.add_argument("--new", "-n", help="Tell the program to create a brand new metadata_table.xlsx (as opposed to updating it with one line (this takes quite a bit longer).", action="store_true")
parser.add_argument("--delete", "-d", help="Tell the program to remove entries with matching genome id numbers (the id numbers would be sent in as a list sparated by '-'s.")
args = parser.parse_args()

engine = create_engine('postgresql://wolf:Halflife3@localhost/BLISTR.db') 
Session = sessionmaker(bind=engine) #need to execute these lines to create a queriable session
session = Session()

workbook_path = 'PATH/TO/ROOT/Database_Inputs/'

#TODO: add flagellen stuff
#TODO: add quality stats and toxin type/group to headers (do this after you get the rest of this program working)
if args.new: #if we are making a new table from scratch
    from blessings import Terminal
    colours = Terminal()
    print colours.cyan + "Creating Metadata Table: " + colours.magenta
    num_lines = session.query(Genome).count() + 1 #plus one for the header
    with tqdm(total=num_lines) as pbar:
        # Create an new Excel file and add a worksheet.
        workbook = xlsxwriter.Workbook(workbook_path+'metadata_table.xlsx')
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({'bold': 1}) #allowing bold wording because why not
    
        headers = ["Check boxes", "Genome name", "Collection date", "Source type", "Source information", "Institution", "Author", "Coding Genes", "ORFs", "Toxin", "BoNT Cassette", "SubType", "bp before toxin", "bp after toxin", "Toxin++ protein figure", "Lat/Long", "Country", "Region"]
        letter = "a"
        for col in headers: #adding all the headers to the file
            worksheet.write(letter.upper()+'1', col, bold) #sets the cell postion (i.e. A1, B1, C1 etc.) and makes the header bold
            letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter]) #wizard one linner that will increase the letter value of letter (i.e 'a' -> 'b')
        pbar.update(1)
        
        g_list = session.query(Genome).order_by(Genome.id.asc()).all() #list of all the genomes in the database
        row = 1 #initializing counter to keep track of row number
        for g in g_list: #filling through all the genomes in the database. g can also be seen as a row in terms of the spreadsheet
            sub_type=str(g.sub_type)
            toxins = session.query(Toxin).filter_by(genome_id=g.id).all() #toxin data that corresponds to the genome in question
            geo = session.query(GeographicLocation).filter_by(genome_id=g.id).first()
            prokka = session.query(Prokka).filter_by(genome_id=g.id).all()
            #The following is just some annoying cleanup incase there are not toxin entries for a genome
            t_meta, cas, bp_before, bp_after = [""]*4 #If there was not toxin input, give these values None
            for tox in toxins:
                if tox != None: 
                   t_meta=t_meta+tox.best_hit+", " #If there was a toxin input, give the outputs their proper values
                   cas=cas+str(tox.orfx_ha)+"; "
                   bp_before=bp_before+str(tox.bp_before)+", "
                   bp_after=bp_after+str(tox.bp_after)+", "
            t_meta, cas, bp_before, bp_after = t_meta.rstrip(", "), cas.rstrip("; "), bp_before.rstrip(", "), bp_after.rstrip(", ") #clean up the last entry
            
            if len(prokka) == 1: #if there is one prokka result
                pro = prokka[0]
                cg = pro.fasta_coding_genes
                ORFs = pro.fasta_ORFs
                if pro.tox_exists:
                    #image = '<a href="images/'+str(g.name)+'.bmp">View</a>'
                    image = "<a href=\"{{ url_for('images', file_name='"+str(g.name)+".bmp') }}\">View</a>"
                else:
                    image = None
            elif len(prokka) > 1: 
                image = ""
                cg = prokka[0].fasta_coding_genes #both these numbers will be base off only the first entry
                ORFs = prokka[0].fasta_ORFs
                counter = 1
                for pro in prokka:            
                    if pro.tox_exists:
                        image = image + "<a href=\"{{ url_for('images', file_name='"+str(g.name)+"."+str(counter)+".bmp') }}\">Toxin"+str(counter)+"  </a>"
                        counter += 1
                    
            else: #incase prokka is NoneType (i.e. in the event prokka is running in the background on BLISTR)
                image = None
                cg = None
                ORFs = None
                
            #Metadata is a list with values that correspond to the values in the list 'headers'
            #The odd replaces on the g.name is to replace things such as type A-B- with A(B) (as it should be). It is done thrice to ensure all '-'s are removed
            metadata = [g.name, g.collection_date, g.clinical_or_environmental, g.source_info, g.institution, g.author, cg, ORFs, t_meta, cas, sub_type, bp_before, bp_after, image,str(geo.lat)[:6]+"/"+str(geo.lng)[:6], geo.country, geo.region.rstrip("\n")]
            letter = "b" #'b' because 'a' is reserved for check boxes
            row += 1 #row starts at 2 and increases from there
            check_box = "<input type=checkbox name=check value="+str(g.id)+">" #the html/jninja required to make a check box, synced to the genome id. All check boxes have the name 'check' which makes for an easy 'select all' box
            worksheet.write("A"+str(row), check_box)
            counter = 0
            for col in metadata:
                if counter == 0: #the following will replace '-'s with the proper brackets should there be need in the genome name
                    if len(col.split(".")) > 1:
                        if len(col.split(".")) == 3 and "-" in col.split(".")[1]: #if there is a '-' in the toxin type (i.e. .A-B-.)
                            col = col.split(".")[0]+"."+col.split(".")[1].replace("-","(",1).replace("-",")",1)+"."+col.split(".")[2]
                        elif len(col.split(".")) == 2 and "-" in col.split(".")[1]:
                            col = col.split(".")[0]+"."+col.split(".")[1].replace("-","(",1).replace("-",")",1)
                    counter += 1
                    
                worksheet.write(letter.upper()+str(row), col)
                letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter]) 
            pbar.update(1)
    print "" + colours.normal
    
    
elif args.delete:   
    #if we are only removing entries from the table
    import os
    from shutil import copyfile
    ids_to_remove = args.delete.split("-")
    copyfile(workbook_path+'metadata_table.xlsx', workbook_path+'old_metadata_table.xlsx') #creating a copy of our old notebook for updating purposes. It will be deleted at the end of this script
    #Input workbook
    wb = load_workbook(filename=workbook_path+'old_metadata_table.xlsx', read_only=True)
    ws = wb['Sheet1'] # ws is now an IterableWorksheet
    
    #Output workbook
    #override the old metadata_table.xlsx file with the new one!    
    workbook = xlsxwriter.Workbook(workbook_path+'metadata_table.xlsx')
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({'bold': 1}) #allowing bold wording because why not    

    ids_to_remove.sort() #sort is a built in function that sorts. It is strange that the interpretor takes this as an undefined entry
        
    #Copy the old workbook ...
    row_num = 0
    for row in ws.rows:
        letter = 'a'
        row_num += 1
        for cell in row:
            if row_num == 1: #for the header row, we want to copy it as is
                worksheet.write(letter.upper()+'1', str(cell.value), bold) #sets the cell postion (i.e. A1, B1, C1 etc.) and makes the header bold
                letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter]) #wizard one linner that will increase the letter value of letter (i.e 'a' -> 'b')
            else:
                if len(ids_to_remove) > 0 and ids_to_remove[0] and letter.lower() == "a" and str(cell.value) == "<input type=checkbox name=check value="+str(ids_to_remove[0])+">": #if the first value is our checkbox with the matching genome id
                    ids_to_remove = ids_to_remove[1:] #remove this entry from our to_update lists (so that the first entry is always the one to be updated)
                    row_num -= 1 #since we are skipping the row, we have to push all the following rows up one
                    break #break to go to the next row (NOT the next column) and not copy anything for this particular row
                    
                else: #normal copying
                    worksheet.write(letter.upper()+str(row_num), str(cell.value))
                    letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter])        

    os.remove(workbook_path+'old_metadata_table.xlsx') #clean up the crap
    
    
else: #if we are only updating the table
    import os
    from shutil import copyfile
    from sqlalchemy.sql.expression import func
    copyfile(workbook_path+'metadata_table.xlsx', workbook_path+'old_metadata_table.xlsx') #creating a copy of our old notebook for updating purposes. It will be deleted at the end of this script
    #Input workbook
    wb = load_workbook(filename=workbook_path+'old_metadata_table.xlsx', read_only=True)
    ws = wb['Sheet1'] # ws is now an IterableWorksheet
    
    #Output workbook
    #override the old metadata_table.xlsx file with the new one!    
    workbook = xlsxwriter.Workbook(workbook_path+'metadata_table.xlsx')
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({'bold': 1}) #allowing bold wording because why not    
    

    #we will update any entries that have the characteristic 'is_analyzed' set to False, then we will set it to True after the addition
    genomes = session.query(Genome).order_by(Genome.id.asc()).all()
    genomes_to_update_ids = [] #initialize list of genome ids of genomes which need changing
    for genome in genomes:
        if not genome.is_analyzed:
            genomes_to_update_ids.append(int(genome.id))
    genomes_to_update_ids.sort() #sort is a built in function that sorts. It is strange that the interpretor takes this as an undefined entry
    num_of_updates = len(genomes_to_update_ids)
    genomes_to_update_metadata = [] #list that will hold lists of all the genomic information of entries to be changed
    for g_id in genomes_to_update_ids:
        g = session.query(Genome).filter_by(id=g_id).first()
        sub_type=str(g.sub_type)
        toxins = session.query(Toxin).filter_by(genome_id=g_id).all() #toxin data that corresponds to the genome in question
        geo = session.query(GeographicLocation).filter_by(genome_id=g_id).first()
        prokka = session.query(Prokka).filter_by(genome_id=g_id).all()
        
        #The following is just some annoying cleanup incase there are not toxin entries for a genome
        t_meta, cas, bp_before, bp_after = [""]*4 #If there was not toxin input, give these values None
        for tox in toxins:
            if tox != None: 
               t_meta=t_meta+tox.best_hit+", " #If there was a toxin input, give the outputs their proper values
               cas=cas+str(tox.orfx_ha)+"; "
               bp_before=bp_before+str(tox.bp_before)+", "
               bp_after=bp_after+str(tox.bp_after)+", "
        t_meta, cas, bp_before, bp_after = t_meta.rstrip(", "), cas.rstrip("; "), bp_before.rstrip(", "), bp_after.rstrip(", ")
           
        if len(prokka) == 1:
            pro = prokka[0]
            cg = pro.fasta_coding_genes
            ORFs = pro.fasta_ORFs
            if pro.tox_exists:
                #image = '<a href="images/'+str(g.name)+'.bmp">View</a>'
                image = "<a href=\"{{ url_for('images', file_name='"+str(g.name)+".bmp') }}\">View</a>"                
            else:
                image = None
        elif len(prokka) > 1:
            image = ""
            cg = prokka[0].fasta_coding_genes #both these numbers will be base off only the first entry
            ORFs = prokka[0].fasta_ORFs
            counter = 1
            for pro in prokka:            
                if pro.tox_exists:
                    image = image + "<a href=\"{{ url_for('images', file_name='"+str(g.name)+"."+str(counter)+".bmp') }}\">Toxin"+str(counter)+"  </a>"
                    counter += 1
                
        else: #incase pro is NoneType
            image = None
            cg = None
            ORFs = None
            
            
        #Metadata is a list with values that correspond to the values in the list 'headers'
        genomes_to_update_metadata.append(["<input type=checkbox name=check value="+str(g.id)+">", g.name, g.collection_date, g.clinical_or_environmental, g.source_info, g.institution, g.author, cg, ORFs, t_meta, cas, sub_type, bp_before, bp_after, image,str(geo.lat)[:6]+"/"+str(geo.lng)[:6], geo.country, geo.region.rstrip("\n")])
    #Now we alctually copy the data to the new file
    #Copy the old workbook ...
    row_num = 0
    for row in ws.rows:
        letter = 'a'
        row_num += 1
        for cell in row:
            if row_num == 1: #for the header row, we want to copy it as is
                worksheet.write(letter.upper()+'1', str(cell.value), bold) #sets the cell postion (i.e. A1, B1, C1 etc.) and makes the header bold
                letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter]) #wizard one linner that will increase the letter value of letter (i.e 'a' -> 'b')
            else:
                if len(genomes_to_update_metadata) > 0 and genomes_to_update_metadata[0] and str(cell.value) in genomes_to_update_metadata[0] and letter.lower() == "a": #if the first value is our checkbox with the matching genome id
                    for data in genomes_to_update_metadata[0]:
                        worksheet.write(letter.upper()+str(row_num), str(data)) #write out all the changed data to the current row
                        letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter])
                    updating_genome = session.query(Genome).filter_by(id=genomes_to_update_ids[0]).first()
                    updating_genome.is_analyzed = True #store the fact that it has been updated
                    session.add(updating_genome)
                    session.commit()
                    genomes_to_update_ids = genomes_to_update_ids[1:] #remove this entry from our to_update lists (so that the first entry is always the one to be updated)
                    genomes_to_update_metadata = genomes_to_update_metadata[1:]
                    break #break to go to the next row (NOT the next column)
                    
                else: #normal copying
                    worksheet.write(letter.upper()+str(row_num), str(cell.value))
                    letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter])
                    
    
    #add in the remaining genomes that have not been covered by the above
    for this_is_just_a_formality in genomes_to_update_ids:
        row_num += 1 #update to the latest row number
        letter = 'a'
        for data in genomes_to_update_metadata[0]:
            worksheet.write(letter.upper()+str(row_num), str(data)) #write out all the changed data to the current row
            letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter])
        updating_genome = session.query(Genome).filter_by(id=genomes_to_update_ids[0]).first()
        updating_genome.is_analyzed = True #store the fact that it has been updated
        session.add(updating_genome)
        session.commit()
        genomes_to_update_ids = genomes_to_update_ids[1:]
        genomes_to_update_metadata = genomes_to_update_metadata[1:]
        

    os.remove(workbook_path+'old_metadata_table.xlsx') #clean up the crap
    
workbook.close()