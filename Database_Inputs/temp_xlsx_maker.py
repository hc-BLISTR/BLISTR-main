# -*- coding: utf-8 -*-
"""
Created on Tue Jul 12 07:15:14 2016

@author: Owen Neuber

Program which generates a .xlsx file (called temp.xlsx) to be stored in the
directory: PATH/TO/ROOT/Database_Inputs/ 
This .xlsx file will hold a table of all the rows which have a matching word
from a search request given by a user. This progam will analyze/copy the rows
from PATH/TO/ROOT/Database_Inputs/metadata_table.xlsx for 
this purpose. 
"""

import xlsxwriter
from openpyxl import load_workbook
import argparse
parser = argparse.ArgumentParser()
parser.add_argument("search", help="Pass in the string which you would like to search for in the database metadata table")
parser.add_argument("--issue", "-i", help="Tell the program to base its search table off of 16S_issues", action="store_true")
args = parser.parse_args()

counter = 0
searches = args.search.replace("-", " ").replace("~","-").lower().split(" and ") #resetting the dashed in entry with spaces

path = 'PATH/TO/ROOT/Database_Inputs/'
toxin_row = 10

for search in searches:
    table = 'temp.xlsx'
    if counter == 0: #counter to make sure we only bother with setting this one time
        table = 'metadata_table.xlsx'
        if args.issue: #tells us that we are making a temp table for the 16S issues page 
            table = '16S_issues.xlsx' #changing variables accordingly
            toxin_row = 9
    #Input workbook
    wb = load_workbook(filename=path+table, read_only=True) #workbook version of open(file, 'r')
    ws = wb['Sheet1'] # ws is now an IterableWorksheet
    
    #Output workbook
    # Create an new Excel file and add a worksheet.
    workbook = xlsxwriter.Workbook(path+'temp.xlsx') #wookbox version of open(file, 'w')
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({'bold': 1}) #allowing bold wording because why not
    
    write_out_row = 1
    counter = 0
    row_num = 0
    for row in ws.rows:
        letter = 'a' #initializing our workbook letter (i.e. A1)
        counter += 1
        row_num += 1
        current_cell_number = 1 #if this is 2, that means we are on the cell holding the genome name. This is of great significance should the user be searching for a group or toxin type
        store = [] #variable to store the current row data in
        add = False #if True, the current row will be added to the outputted excel file
        for cell in row:
            if counter == 1: #for the header row, we want to copy it as is
                worksheet.write(letter.upper()+'1', cell.value, bold) #sets the cell postion (i.e. A1, B1, C1 etc.) and makes the header bold
                letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter]) #wizard one linner that will increase the letter value of letter (i.e 'a' -> 'b')
            else:
                store.append(str(cell.value)) #store everything in the row in this list (all to be written out to our temp workbook should a matching thing be found)
                
                try:
                    int(cell.value) #check for integer values and excude any of those from the toxin type uniqueness
                    is_int = True
                except:
                    is_int = False
                #this below line is what alows me to determine if they are searching for a toxin type or something else
                if search=="a" or search=="b" or search=="c" or search=="c/d" or search=="d/c" or search=="d" or search=="e" or search=="f" or search=="g": #in case they are searching for, 'a' (as in 'type a'), or for group 1. The replace to to ensure that dashes don't skew the character count on the genome type
                    if current_cell_number == toxin_row and cell.value != None and cell.value.lower() != "none": #if we are on the genome name cell, the following becomes relavent
                        g_type_list = cell.value.split(", ")
                        for g_type in g_type_list:
                            if search.upper() in g_type.upper():
                                add = True #if we set add to True at any point, all the information from store will be added to our temp file
                                break
                
                else: #If they are not looking for a group number or toxin type, a simple in will suffice
                    if str(search).upper() in str(cell.value).upper():
                        add = True
            current_cell_number += 1
            
        if add: #now that the whole line has been parsed
            write_out_row += 1 #increase the row number to which we will write out
            for item in store: #add the whole line to the excel file       
                worksheet.write(letter.upper()+str(write_out_row), item) #sets the cell postion (i.e. A1, B1, C1 etc.) and makes the header bold
                letter = ''.join([chr(((ord(x)-ord('a')+1)%26)+ord('a')) for x in letter])
                
    workbook.close()